from pathlib import Path
from openpyxl import load_workbook

from renderer import create_card
from config import load_config   # you already have config.py

# -----------------------
# Paths
# -----------------------
BASE_DIR = Path(__file__).resolve().parent.parent

excel_file = BASE_DIR / "excel" / "cards.xlsx"
output_dir = BASE_DIR / "output"

# -----------------------
# Load config
# -----------------------
config = load_config(BASE_DIR / "config.json")

# -----------------------
# Ensure output folder exists
# -----------------------
output_dir.mkdir(exist_ok=True)

# -----------------------
# Load Excel file
# -----------------------
wb = load_workbook(excel_file)
ws = wb.active

print("🚀 Starting card generation...\n")

# -----------------------
# Main loop
# -----------------------
for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):

    # Defensive parsing (prevents crash on empty rows)
    if not row or not row[0] or not row[1]:
        print(f"⚠️ Skipping row {index} (missing data)")
        continue

    question = str(row[0]).strip()
    answer = str(row[1]).strip()

    output_file = output_dir / f"Card_{index:03}.png"

    # -----------------------
    # Render card
    # -----------------------
    create_card(
        question=question,
        answer=answer,
        output_file=output_file,
        config=config   # future-proof hook
    )

    print(f"✅ Generated: {output_file.name}")

print("\n🎉 All cards generated successfully!")