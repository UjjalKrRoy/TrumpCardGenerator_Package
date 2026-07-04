from pathlib import Path
from openpyxl import load_workbook

from src.renderer import create_card


# =========================================================
# CARD GENERATOR ENGINE
# =========================================================
class CardGenerator:
    """
    Batch generator for card images from Excel.

    Responsibilities:
    - Read Excel safely
    - Validate rows
    - Call renderer
    - Report progress
    """

    def __init__(
        self,
        config,
        template_path,
        font_path,
        excel_path,
        output_dir,
        logger=None,
    ):
        self.config = config
        self.template_path = Path(template_path)
        self.font_path = font_path
        self.excel_path = Path(excel_path)
        self.output_dir = Path(output_dir)
        self.logger = logger

    # =====================================================
    # LOGGING
    # =====================================================
    def log(self, msg):
        if self.logger:
            self.logger(msg)

    # =====================================================
    # VALIDATION
    # =====================================================
    def validate_inputs(self):
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {self.template_path}")

        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

        self.output_dir.mkdir(parents=True, exist_ok=True)

    # =====================================================
    # MAIN GENERATION FUNCTION
    # =====================================================
    def generate(self):
        self.validate_inputs()

        workbook = load_workbook(self.excel_path)
        sheet = workbook.active

        generated = 0
        skipped = 0

        # =================================================
        # READ EXCEL ROWS
        # =================================================
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=1,
        ):
            try:
                # ---------------- VALIDATION ----------------
                if not row or len(row) < 2:
                    self.log(f"Skipping row {row_index}: invalid structure")
                    skipped += 1
                    continue

                question, answer = row[0], row[1]

                if not question or not answer:
                    self.log(f"Skipping row {row_index}: missing data")
                    skipped += 1
                    continue

                # ---------------- OUTPUT PATH ----------------
                output_file = self.output_dir / f"Card_{row_index:03}.png"

                # ---------------- RENDER ----------------
                create_card(
                    question=str(question).strip(),
                    answer=str(answer).strip(),
                    output_file=output_file,
                    config=self.config,
                    template_path=self.template_path,
                    font_path=self.font_path,
                )

                generated += 1
                self.log(f"Generated: {output_file.name}")

            except Exception as e:
                skipped += 1
                self.log(f"Error row {row_index}: {str(e)}")

        # =================================================
        # SUMMARY
        # =================================================
        self.log(f"\nDone → Generated: {generated}, Skipped: {skipped}")

        return {
            "generated": generated,
            "skipped": skipped,
        }