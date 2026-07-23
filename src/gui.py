import webbrowser
from pathlib import Path
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, colorchooser
from PIL import ImageTk
from openpyxl import load_workbook
import csv
from docx import Document
from copy import deepcopy

from matplotlib import font_manager

from src.config import load_config
from src.renderer import create_card
from src.updater import check_for_update, download_and_replace


# =========================================================
# SYSTEM FONTS
# =========================================================
def get_system_fonts():
    fonts = {}

    for f in font_manager.fontManager.ttflist:

        name = f.name
        filename = Path(f.fname).stem.lower()

        # If this is the regular font (not bold/italic), always use it.
        if filename.endswith(("bd", "bi", "i")) or "bold" in filename or "italic" in filename:
            if name not in fonts:
                fonts[name] = f.fname
        else:
            # Regular font replaces any previous variant.
            fonts[name] = f.fname

    return fonts

# =========================================================
# MAIN APP
# =========================================================
class CardGeneratorGUI:
    def create_scrollable_tab(self):
        container = ttk.Frame(self.notebook)

        canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(
            container,
            orient="vertical",
            command=canvas.yview
        )

        scroll_frame = ttk.Frame(canvas)

        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        window = canvas.create_window(
            (0, 0),
            window=scroll_frame,
            anchor="nw"
        )

        canvas.bind(
            "<Configure>",
            lambda e: canvas.itemconfigure(
                window,
                width=e.width
            )
        )

        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def _wheel(event):
            canvas.yview_scroll(int(-event.delta / 120), "units")

        scroll_frame.bind(
            "<Enter>",
            lambda e: canvas.bind_all("<MouseWheel>", _wheel)
        )

        scroll_frame.bind(
            "<Leave>",
            lambda e: canvas.unbind_all("<MouseWheel>")
        )

        return container, scroll_frame
    def __init__(self, root):
        self.root = root

        base_dir = Path(__file__).resolve().parent.parent

        self.root.title("Trump Card Generator v1.0.28")

        icon_path = base_dir / "icon.ico"
        if icon_path.exists():
            self.root.iconbitmap(icon_path)

        self.config_obj = load_config(base_dir / "config.json")
        
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()

        width = int(screen_w * 0.9)
        height = int(screen_h * 0.9)

        x = (screen_w - width) // 2
        y = (screen_h - height) // 2

        self.root.geometry(f"{width}x{height}+{x}+{y}")
        self.root.minsize(1100, 700)

        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)

        base_dir = Path(__file__).resolve().parent.parent
        self.config_path = base_dir / "config.json"
        self.config_obj = load_config(self.config_path)

        # ---------------- FILES ----------------
        self.green_template_path = None
        self.red_template_path = None

        self.excel_path = None
        self.output_dir = None

        # ---------------- FONTS ----------------
        self.font_map = get_system_fonts()
        self.font_names = sorted(self.font_map.keys())
        self.font_var = tk.StringVar(value="Arial")

        # ---------------- STATE ----------------
        self.preview_photo = None
        self.base_boxes = deepcopy(self.config_obj.boxes)

        # ---------------- STYLE ----------------
        q = self.config_obj.get_style("question")
        a = self.config_obj.get_style("answer")

        self.q_max = tk.IntVar(value=q.get("font_max", 32))
        self.q_min = tk.IntVar(value=q.get("font_min", 18))
        self.a_max = tk.IntVar(value=a.get("font_max", 28))
        self.a_min = tk.IntVar(value=a.get("font_min", 20))

        # Green Card Colors
        self.green_q_fill = "#FFFFFF"
        self.green_a_fill = "#FFFFFF"

        self.green_q_outline = "#000000"
        self.green_a_outline = "#000000"

        # Red Card Colors
        self.red_q_fill = "#FFFF00"
        self.red_a_fill = "#FFFF00"

        self.red_q_outline = "#000000"
        self.red_a_outline = "#000000"

        # Subject & Stage Colors

        self.green_subject_fill = "#FFFFFF"
        self.green_subject_outline = "#000000"

        self.green_stage_fill = "#FFFFFF"
        self.green_stage_outline = "#000000"

        self.red_subject_fill = "#FFFF00"
        self.red_subject_outline = "#000000"

        self.red_stage_fill = "#FFFF00"
        self.red_stage_outline = "#000000"

        self.q_offset = tk.IntVar(value=0)
        self.a_offset = tk.IntVar(value=0)

        # Question box resize
        self.q_left = tk.IntVar(value=0)
        self.q_right = tk.IntVar(value=0)
        self.q_top = tk.IntVar(value=0)
        self.q_bottom = tk.IntVar(value=0)

        # Answer box resize
        self.a_left = tk.IntVar(value=0)
        self.a_right = tk.IntVar(value=0)
        self.a_top = tk.IntVar(value=0)
        self.a_bottom = tk.IntVar(value=0)

        self.q_margin_left = tk.IntVar(value=20)
        self.q_margin_right = tk.IntVar(value=20)
        self.q_margin_top = tk.IntVar(value=10)
        self.q_margin_bottom = tk.IntVar(value=10)

        self.a_margin_left = tk.IntVar(value=20)
        self.a_margin_right = tk.IntVar(value=20)
        self.a_margin_top = tk.IntVar(value=10)
        self.a_margin_bottom = tk.IntVar(value=10)

        # Excel row range
        self.start_row = tk.IntVar(value=2)
        self.end_row = tk.IntVar(value=0)   # 0 = last row

        # =====================================================
        # SUBJECT & STAGE
        # =====================================================

        # Values entered by the user
        self.subject_var = tk.StringVar()
        self.stage_var = tk.StringVar()

        # Fonts
        self.subject_font_var = tk.StringVar(value="Arial")
        self.stage_font_var = tk.StringVar(value="Arial")

        # Font sizes
        self.subject_font_size = tk.IntVar(value=24)
        self.stage_font_size = tk.IntVar(value=24)

        # Bold
        self.subject_bold = tk.BooleanVar(value=False)
        self.stage_bold = tk.BooleanVar(value=False)

        # Italic
        self.subject_italic = tk.BooleanVar(value=False)
        self.stage_italic = tk.BooleanVar(value=False)

        # Position Adjustment
        self.subject_left = tk.IntVar(value=0)
        self.subject_right = tk.IntVar(value=0)
        self.subject_up = tk.IntVar(value=0)
        self.subject_down = tk.IntVar(value=0)

        self.stage_left = tk.IntVar(value=0)
        self.stage_right = tk.IntVar(value=0)
        self.stage_up = tk.IntVar(value=0)
        self.stage_down = tk.IntVar(value=0)

        # ==========================
        # Main Frame
        # ==========================

        self.main_frame = ttk.Frame(self.root)
        self.main_frame.grid(row=0,column=0,sticky="nsew")

        self.main_frame.columnconfigure(0, weight=1, minsize=340)
        self.main_frame.columnconfigure(1, weight=3)
        self.main_frame.rowconfigure(0,weight=1)

        self.left_panel = ttk.Frame(
            self.main_frame,
            padding=10,
            width=360
        )

        self.left_panel.grid(
            row=0,
            column=0,
            sticky="nsew"
        )

        self.notebook = ttk.Notebook(
            self.left_panel
        )

        self.notebook.pack(
            fill="both",
            expand=True,
            padx=2,
            pady=2
        )

        self.left_panel.pack_propagate(False)

        project_container, self.project_tab = self.create_scrollable_tab()
        style_container, self.style_tab = self.create_scrollable_tab()
        layout_container, self.layout_tab = self.create_scrollable_tab()
        color_container, self.color_tab = self.create_scrollable_tab()
        help_container, self.help_tab = self.create_scrollable_tab()

        self.notebook.add(
            project_container,
            text="Project"
        )

        self.notebook.add(
            style_container,
            text="Typography"
        )

        self.notebook.add(
            layout_container,
            text="Layout"
        )

        self.notebook.add(
            color_container,
            text="Colors"
        )

        self.notebook.add(
            help_container,
            text="Help"
        )

        self.right_panel = ttk.Frame(
            self.main_frame,
            padding=15
        )

        self.right_panel.grid(
            row=0,
            column=1,
            sticky="nsew"
        )

        self.right_panel.rowconfigure(0,weight=1)
        self.right_panel.rowconfigure(1,weight=0)
        self.right_panel.columnconfigure(0, weight=1)

       
        # Auto-update preview when values change
        for var in (

            self.q_max,
            self.q_min,

            self.a_max,
            self.a_min,

            self.q_offset,
            self.a_offset,

            self.q_margin_left,
            self.q_margin_right,
            self.q_margin_top,
            self.q_margin_bottom,

            self.a_margin_left,
            self.a_margin_right,
            self.a_margin_top,
            self.a_margin_bottom,

            self.q_left,
            self.q_right,
            self.q_top,
            self.q_bottom,

            self.a_left,
            self.a_right,
            self.a_top,
            self.a_bottom,

            self.subject_var,
            self.stage_var,

            self.subject_font_var,
            self.stage_font_var,

            self.subject_font_size,
            self.stage_font_size,

            self.subject_bold,
            self.subject_italic,

            self.stage_bold,
            self.stage_italic,

            self.subject_left,
            self.subject_right,
            self.subject_up,
            self.subject_down,

            self.stage_left,
            self.stage_right,
            self.stage_up,
            self.stage_down,

        ):
            var.trace_add("write", lambda *args: self.update_preview())
    
        # =====================================================
        # BUILD UI SECTIONS
        # =====================================================
        self.build_project_section()
        self.build_font_section()
        self.build_style_section()
        self.build_layout_section()
        self.build_color_section()
        self.build_help_section()
        self.build_actions_section()

        # =====================================================
        # PREVIEW PANEL
        # =====================================================

        preview_frame = ttk.LabelFrame(
            self.right_panel,
            text="Card Preview"
        )

        preview_frame.grid(
            row=0,
            column=0,
            sticky="nsew",
            padx=10,
            pady=10
        )

        preview_frame.rowconfigure(0, weight=1)
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.columnconfigure(1, weight=1)

        green_frame = ttk.LabelFrame(
            preview_frame,
            text="Green Card"
        )
        green_frame.grid(
            row=0,
            column=0,
            sticky="nsew",
            padx=5,
            pady=5
        )

        red_frame = ttk.LabelFrame(
            preview_frame,
            text="Red Card"
        )
        red_frame.grid(
            row=0,
            column=1,
            sticky="nsew",
            padx=5,
            pady=5
        )

        self.green_preview = ttk.Label(green_frame)
        self.green_preview.pack(fill="both", expand=True)

        self.red_preview = ttk.Label(red_frame)
        self.red_preview.pack(fill="both", expand=True)

        self.green_photo = None
        self.red_photo = None

        log_frame = ttk.LabelFrame(
            self.right_panel,
            text="Generation Log"
        )

        log_frame.grid(
            row=1,
            column=0,
            sticky="ew",
            pady=(10,0)
        )

        self.log_box = scrolledtext.ScrolledText(
            log_frame,
            height=6
        )

        self.log_box.pack(
            fill="both",
            expand=True
        )
        self.status = tk.StringVar(value="Ready")

        ttk.Label(
            self.root,
            textvariable=self.status,
            relief="sunken",
            anchor="w"
        ).grid(
            row=1,
            column=0,
            sticky="ew"
        )
    # =====================================================
    # UI SECTIONS
    # =====================================================

    def section(self,parent,title):

        frame=ttk.LabelFrame(
            parent,
            text=title,
            padding=10
        )

        frame.pack(fill="x", pady=10)

        return frame

    # ---------------- PROJECT ----------------
    def build_project_section(self):
        f = self.section(
            self.project_tab,
            "Project"
        )

        ttk.Button(f, text="Select Green Card", command=self.select_green_template).pack(fill="x")
        self.green_template_label = ttk.Label(f, text="Not Selected", foreground="gray", wraplength=300)
        self.green_template_label.pack(anchor="w", pady=(0,5))
        ttk.Button(f, text="Select Red Card", command=self.select_red_template).pack(fill="x")
        self.red_template_label = ttk.Label(f, text="Not Selected", foreground="gray", wraplength=300)
        self.red_template_label.pack(anchor="w", pady=(0,5))
        ttk.Button(f, text="Select Data File", command=self.select_excel).pack(fill="x", pady=2)
        self.excel_label = ttk.Label(f, text="Not Selected", foreground="gray", wraplength=300)
        self.excel_label.pack(anchor="w", pady=(0,5))
        ttk.Button(f, text="Select Output Folder", command=self.select_output).pack(fill="x")
        self.output_label = ttk.Label(f, text="Not Selected", foreground="gray", wraplength=300)
        self.output_label.pack(anchor="w", pady=(0,5))
        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=8)
        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=8)

        # =====================================================
        # SUBJECT
        # =====================================================

        ttk.Label(f, text="Subject").pack(anchor="w")
        ttk.Entry(f, textvariable=self.subject_var).pack(fill="x")

        # =====================================================
        # STAGE
        # =====================================================

        ttk.Label(f, text="Stage").pack(anchor="w", pady=(8,0))
        ttk.Entry(f, textvariable=self.stage_var).pack(fill="x")
        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=8)

        ttk.Label(f, text="Start Row").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=1,
            to=100000,
            textvariable=self.start_row
        ).pack(fill="x")

        ttk.Label(f, text="End Row (0 = Last Row)").pack(anchor="w")
        ttk.Label(
            f,
            text="0 = Generate till last row",
            foreground="gray"
        ).pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=0,
            to=100000,
            textvariable=self.end_row
        ).pack(fill="x")

    # ---------------- FONT ----------------
    def build_font_section(self):
        f = self.section(
            self.style_tab,
            "Typography"
        )

        ttk.Label(f, text="Font").pack(anchor="w")

        self.font_dropdown = ttk.Combobox(
            f,
            textvariable=self.font_var,
            values=self.font_names,
            state="readonly"
        )
        self.font_dropdown.pack(fill="x")

        self.font_var.trace_add(
            "write",
            lambda *args: self.update_preview()
        )

    # ---------------- STYLE ----------------
    def build_style_section(self):
        f = self.section(
            self.style_tab,
            "Font Size"
        )

        ttk.Label(f, text="Question Max").pack(anchor="w")
        ttk.Spinbox(f, from_=8, to=120, textvariable=self.q_max).pack(fill="x")

        ttk.Label(f, text="Question Min").pack(anchor="w")
        ttk.Spinbox(f, from_=8, to=120, textvariable=self.q_min).pack(fill="x")

        ttk.Label(f, text="Answer Max").pack(anchor="w")
        ttk.Spinbox(f, from_=8, to=120, textvariable=self.a_max).pack(fill="x")

        ttk.Label(f, text="Answer Min").pack(anchor="w")
        ttk.Spinbox(f, from_=8, to=120, textvariable=self.a_min).pack(fill="x")

        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=10)
        ttk.Label(f, text="Subject",font=("Segoe UI",10,"bold")).pack(anchor="w")
        ttk.Label(f, text="Font").pack(anchor="w")
        ttk.Combobox(f, textvariable=self.subject_font_var, values=self.font_names, state="readonly").pack(fill="x")
        ttk.Label(f, text="Font Size").pack(anchor="w")
        ttk.Spinbox(f, from_=8, to=120, textvariable=self.subject_font_size).pack(fill="x")
        ttk.Checkbutton(f, text="Bold", variable=self.subject_bold).pack(anchor="w")
        ttk.Checkbutton(f, text="Italic", variable=self.subject_italic).pack(anchor="w")

        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=10)
        ttk.Label(f, text="Stage", font=("Segoe UI",10,"bold")).pack(anchor="w")
        ttk.Label(f, text="Font").pack(anchor="w")
        ttk.Combobox(f, textvariable=self.stage_font_var, values=self.font_names, state="readonly").pack(fill="x")
        ttk.Label(f, text="Font Size").pack(anchor="w")
        ttk.Spinbox(f, from_=8, to=120, textvariable=self.stage_font_size).pack(fill="x")
        ttk.Checkbutton(f, text="Bold", variable=self.stage_bold).pack(anchor="w")
        ttk.Checkbutton(f, text="Italic", variable=self.stage_italic).pack(anchor="w")

    # ---------------- LAYOUT ----------------
    def build_layout_section(self):
        f = self.section(
            self.layout_tab,
            "Layout"
        )

        # ===========================
        # Question
        # ===========================
        ttk.Label(f, text="Question Y Offset").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-200,
            to=200,
            textvariable=self.q_offset
        ).pack(fill="x")

        ttk.Label(f, text="Question Left Margin").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=0,
            to=100,
            textvariable=self.q_margin_left
        ).pack(fill="x")

        ttk.Label(f, text="Question Right Margin").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=0,
            to=100,
            textvariable=self.q_margin_right
        ).pack(fill="x")

        ttk.Label(f, text="Question Top Margin").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=0,
            to=100,
            textvariable=self.q_margin_top
        ).pack(fill="x")

        ttk.Label(f, text="Question Bottom Margin").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=0,
            to=100,
            textvariable=self.q_margin_bottom
        ).pack(fill="x")

        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=8)

        ttk.Label(f, text="Question Box Left").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.q_left
        ).pack(fill="x")

        ttk.Label(f, text="Question Box Right").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.q_right
        ).pack(fill="x")

        ttk.Label(f, text="Question Box Top").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.q_top
        ).pack(fill="x")

        ttk.Label(f, text="Question Box Bottom").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.q_bottom
        ).pack(fill="x")
        
        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=8)

        # ===========================
        # Answer
        # ===========================
        ttk.Label(f, text="Answer Y Offset").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-200,
            to=200,
            textvariable=self.a_offset
        ).pack(fill="x")

        ttk.Label(f, text="Answer Left Margin").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=0,
            to=100,
            textvariable=self.a_margin_left
        ).pack(fill="x")

        ttk.Label(f, text="Answer Right Margin").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=0,
            to=100,
            textvariable=self.a_margin_right
        ).pack(fill="x")

        ttk.Label(f, text="Answer Top Margin").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=0,
            to=100,
            textvariable=self.a_margin_top
        ).pack(fill="x")

        ttk.Label(f, text="Answer Bottom Margin").pack(anchor="w")
        
        ttk.Spinbox(
            f,
            from_=0,
            to=100,
            textvariable=self.a_margin_bottom
        ).pack(fill="x")

        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=8)

        ttk.Label(f, text="Answer Box Left").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.a_left
        ).pack(fill="x")

        ttk.Label(f, text="Answer Box Right").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.a_right
        ).pack(fill="x")

        ttk.Label(f, text="Answer Box Top").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.a_top
        ).pack(fill="x")

        ttk.Label(f, text="Answer Box Bottom").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.a_bottom
        ).pack(fill="x")

        # =====================================================
        # SUBJECT
        # =====================================================

        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=10)

        ttk.Label(
            f,
            text="Subject",
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w")

        ttk.Label(f, text="Move Left").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.subject_left
        ).pack(fill="x")

        ttk.Label(f, text="Move Right").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.subject_right
        ).pack(fill="x")

        ttk.Label(f, text="Move Up").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.subject_up
        ).pack(fill="x")

        ttk.Label(f, text="Move Down").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.subject_down
        ).pack(fill="x")


        # =====================================================
        # STAGE
        # =====================================================

        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=10)

        ttk.Label(
            f,
            text="Stage",
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w")

        ttk.Label(f, text="Move Left").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.stage_left
        ).pack(fill="x")

        ttk.Label(f, text="Move Right").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-300,
            to=300,
            textvariable=self.stage_right
        ).pack(fill="x")

        ttk.Label(f, text="Move Up").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-1000,
            to=1000,
            textvariable=self.stage_up
        ).pack(fill="x")

        ttk.Label(f, text="Move Down").pack(anchor="w")
        ttk.Spinbox(
            f,
            from_=-1000,
            to=1000,
            textvariable=self.stage_down
        ).pack(fill="x")
   
    # ---------------- COLORS ----------------
    def build_color_section(self):

        green = self.section(self.color_tab, "Green Card")

        ttk.Button(green, text="Question Color",
                command=self.pick_green_q_fill).pack(fill="x")

        ttk.Button(green, text="Answer Color",
                command=self.pick_green_a_fill).pack(fill="x")

        ttk.Button(green, text="Question Outline",
                command=self.pick_green_q_outline).pack(fill="x")

        ttk.Button(green, text="Answer Outline",
                command=self.pick_green_a_outline).pack(fill="x")
        
        ttk.Button(
            green,
            text="Subject Color",
            command=self.pick_green_subject_fill
        ).pack(fill="x")

        ttk.Button(
            green,
            text="Subject Outline",
            command=self.pick_green_subject_outline
        ).pack(fill="x")

        ttk.Button(
            green,
            text="Stage Color",
            command=self.pick_green_stage_fill
        ).pack(fill="x")

        ttk.Button(
            green,
            text="Stage Outline",
            command=self.pick_green_stage_outline
        ).pack(fill="x")


        red = self.section(self.color_tab, "Red Card")

        ttk.Button(red, text="Question Color",
                command=self.pick_red_q_fill).pack(fill="x")

        ttk.Button(red, text="Answer Color",
                command=self.pick_red_a_fill).pack(fill="x")

        ttk.Button(red, text="Question Outline",
                command=self.pick_red_q_outline).pack(fill="x")

        ttk.Button(red, text="Answer Outline",
                command=self.pick_red_a_outline).pack(fill="x")
        
        ttk.Button(
            red,
            text="Subject Color",
            command=self.pick_red_subject_fill
        ).pack(fill="x")

        ttk.Button(
            red,
            text="Subject Outline",
            command=self.pick_red_subject_outline
        ).pack(fill="x")

        ttk.Button(
            red,
            text="Stage Color",
            command=self.pick_red_stage_fill
        ).pack(fill="x")

        ttk.Button(
            red,
            text="Stage Outline",
            command=self.pick_red_stage_outline
        ).pack(fill="x")
    # ---------------- HELP ----------------
    def build_help_section(self):

        f = self.section(
            self.help_tab,
            "Software Update"
        )

        ttk.Label(
            f,
            text="Check whether a newer version is available."
        ).pack(anchor="w", pady=(0, 10))

        ttk.Button(
            f,
            text="Check for Updates",
            command=self.check_updates
        ).pack(fill="x")

        ttk.Separator(
            f,
            orient="horizontal"
        ).pack(fill="x", pady=15)

        ttk.Label(
            f,
            text="Application Information"
        ).pack(anchor="w", pady=(0, 10))

        ttk.Button(
            f,
            text="About Trump Card Generator",
            command=self.show_about
        ).pack(fill="x")
    # ---------------- ACTIONS ----------------
    def build_actions_section(self):
        f = self.section(
            self.project_tab,
            "Actions"
        )

        ttk.Button(
            f,
            text="Preview",
            command=self.update_preview
        ).pack(fill="x")

        ttk.Button(
            f,
            text="Reset Settings",
            command=self.reset_settings
        ).pack(fill="x", pady=5)

        ttk.Button(
            f,
            text="Generate Cards",
            command=self.start_generation
        ).pack(fill="x")

    # =====================================================
    # FILE PICKERS
    # =====================================================

    def select_green_template(self):
        f = filedialog.askopenfilename(
            filetypes=[("Images", "*.png *.jpg *.jpeg")]
        )

        if f:
            self.green_template_path = Path(f)

            self.green_template_label.config(
                text=self.green_template_path.name,
                foreground="green"
            )

            self.status.set(
                f"Green Template: {self.green_template_path.name}"
            )

            self.update_preview()
    
    def select_red_template(self):
        f = filedialog.askopenfilename(
            filetypes=[("Images", "*.png *.jpg *.jpeg")]
        )

        if f:
            self.red_template_path = Path(f)

            self.red_template_label.config(
                text=self.red_template_path.name,
                foreground="green"
            )

            self.status.set(
                f"Red Template: {self.red_template_path.name}"
            )

            self.update_preview()

    def select_excel(self):
        f = filedialog.askopenfilename(
            filetypes=[
                ("Supported Files", "*.xlsx *.csv *.txt *.docx"),
                ("Excel", "*.xlsx"),
                ("CSV", "*.csv"),
                ("Text", "*.txt"),
                ("Word", "*.docx"),
            ]
        )

        if f:
            self.excel_path = Path(f)

            self.excel_label.config(
                text=self.excel_path.name,
                foreground="green"
            )

            self.status.set(
                f"Data File: {self.excel_path.name}"
            )

            self.update_preview()

    def select_output(self):
        f = filedialog.askdirectory()

        if f:
            self.output_dir = Path(f)

            self.output_label.config(
                text=str(self.output_dir),
                foreground="green"
            )

            self.status.set(
                f"Output: {self.output_dir}"
            )

            self.update_preview()

    # =====================================================
    # COLORS
    # =====================================================
    
    def pick_green_q_fill(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.green_q_fill = c
            self.update_preview()

    def pick_green_a_fill(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.green_a_fill = c
            self.update_preview()

    def pick_green_q_outline(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.green_q_outline = c
            self.update_preview()

    def pick_green_a_outline(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.green_a_outline = c
            self.update_preview()

    def pick_red_q_fill(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.red_q_fill = c
            self.update_preview()

    def pick_red_a_fill(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.red_a_fill = c
            self.update_preview()

    def pick_red_q_outline(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.red_q_outline = c
            self.update_preview()

    def pick_red_a_outline(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.red_a_outline = c
            self.update_preview()
    
    def pick_green_subject_fill(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.green_subject_fill = c
            self.update_preview()

    def pick_green_subject_outline(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.green_subject_outline = c
            self.update_preview()

    def pick_green_stage_fill(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.green_stage_fill = c
            self.update_preview()

    def pick_green_stage_outline(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.green_stage_outline = c
            self.update_preview()

    def pick_red_subject_fill(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.red_subject_fill = c
            self.update_preview()

    def pick_red_subject_outline(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.red_subject_outline = c
            self.update_preview()

    def pick_red_stage_fill(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.red_stage_fill = c
            self.update_preview()

    def pick_red_stage_outline(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.red_stage_outline = c
            self.update_preview()

    # =====================================================
    # PREVIEW
    # =====================================================

    def update_preview(self):

        config = deepcopy(self.config_obj)

        q = self.base_boxes["question"]
        a = self.base_boxes["answer"]

        config.boxes["question"] = [
            q[0] + self.q_left.get(),
            q[1] + self.q_top.get() + self.q_offset.get(),
            q[2] + self.q_right.get(),
            q[3] + self.q_bottom.get(),
        ]
        config.boxes["answer"] = [
            a[0] + self.a_left.get(),
            a[1] + self.a_offset.get() + self.a_top.get(),
            a[2] + self.a_right.get(),
            a[3] + self.a_bottom.get(),
        ]

        subject_box = self.base_boxes["subject"]

        config.boxes["subject"] = [

            subject_box[0]
            - self.subject_left.get()
            + self.subject_right.get(),

            subject_box[1]
            - self.subject_up.get()
            + self.subject_down.get(),

            subject_box[2],
            subject_box[3],
        ]

        stage_box = self.base_boxes["stage"]

        config.boxes["stage"] = [

            stage_box[0]
            - self.stage_left.get()
            + self.stage_right.get(),

            stage_box[1]
            - self.stage_up.get()
            + self.stage_down.get(),

            stage_box[2],
            stage_box[3],
        ]

        config.style["question"]["font_max"] = self.q_max.get()
        config.style["question"]["font_min"] = self.q_min.get()
        config.style["answer"]["font_max"] = self.a_max.get()
        config.style["answer"]["font_min"] = self.a_min.get()

        # Question margins
        config.style["question"]["margin_left"] = self.q_margin_left.get()
        config.style["question"]["margin_right"] = self.q_margin_right.get()
        config.style["question"]["margin_top"] = self.q_margin_top.get()
        config.style["question"]["margin_bottom"] = self.q_margin_bottom.get()

        # Answer margins
        config.style["answer"]["margin_left"] = self.a_margin_left.get()
        config.style["answer"]["margin_right"] = self.a_margin_right.get()
        config.style["answer"]["margin_top"] = self.a_margin_top.get()
        config.style["answer"]["margin_bottom"] = self.a_margin_bottom.get()

        # Subject typography
        config.style["subject"]["font_size"] = self.subject_font_size.get()
        config.style["subject"]["bold"] = self.subject_bold.get()
        config.style["subject"]["italic"] = self.subject_italic.get()

        # Stage typography
        config.style["stage"]["font_size"] = self.stage_font_size.get()
        config.style["stage"]["bold"] = self.stage_bold.get()
        config.style["stage"]["italic"] = self.stage_italic.get()

        self.green_preview.update_idletasks()

        w = max(self.green_preview.winfo_width() - 20, 300)
        h = max(self.green_preview.winfo_height() - 20, 400)

        # ---------------- GREEN PREVIEW ----------------
        if self.green_template_path:

            green_cfg = deepcopy(config)

            green_cfg.style["question"]["fill"] = self.green_q_fill
            green_cfg.style["answer"]["fill"] = self.green_a_fill
            green_cfg.style["question"]["stroke_fill"] = self.green_q_outline
            green_cfg.style["answer"]["stroke_fill"] = self.green_a_outline
            green_cfg.style["subject"]["fill"] = self.green_subject_fill
            green_cfg.style["subject"]["stroke_fill"] = self.green_subject_outline
            green_cfg.style["stage"]["fill"] = self.green_stage_fill
            green_cfg.style["stage"]["stroke_fill"] = self.green_stage_outline

            green = create_card(
                question="Sample Question",
                answer="Sample Answer",

                subject=self.subject_var.get(),
                stage=self.stage_var.get(),

                subject_font_path=self.font_map.get(
                    self.subject_font_var.get(),
                    "Arial"
                ),

                stage_font_path=self.font_map.get(
                    self.stage_font_var.get(),
                    "Arial"
                ),

                output_file=None,
                config=green_cfg,
                template_path=self.green_template_path,
                font_path=self.font_map.get(
                    self.font_var.get(),
                    "Arial"
                ),
                preview=True,
            )

            green.thumbnail((w, h))

            self.green_photo = ImageTk.PhotoImage(green)
            self.green_preview.configure(image=self.green_photo)

        else:
            self.green_preview.configure(image="")
            self.green_photo = None


        # ---------------- RED PREVIEW ----------------
        if self.red_template_path:

            red_cfg = deepcopy(config)

            red_cfg.style["question"]["fill"] = self.red_q_fill
            red_cfg.style["answer"]["fill"] = self.red_a_fill
            red_cfg.style["question"]["stroke_fill"] = self.red_q_outline
            red_cfg.style["answer"]["stroke_fill"] = self.red_a_outline
            red_cfg.style["subject"]["fill"] = self.red_subject_fill
            red_cfg.style["subject"]["stroke_fill"] = self.red_subject_outline
            red_cfg.style["stage"]["fill"] = self.red_stage_fill
            red_cfg.style["stage"]["stroke_fill"] = self.red_stage_outline

            red = create_card(
                question="Sample Question",
                answer="Sample Answer",

                subject=self.subject_var.get(),
                stage=self.stage_var.get(),

                subject_font_path=self.font_map.get(
                    self.subject_font_var.get(),
                    "Arial"
                ),

                stage_font_path=self.font_map.get(
                    self.stage_font_var.get(),
                    "Arial"
                ),

                output_file=None,
                config=red_cfg,
                template_path=self.red_template_path,
                font_path=self.font_map.get(
                    self.font_var.get(),
                    "Arial"
                ),
                preview=True,
            )

            red.thumbnail((w, h))

            self.red_photo = ImageTk.PhotoImage(red)
            self.red_preview.configure(image=self.red_photo)

        else:
            self.red_preview.configure(image="")
            self.red_photo = None

    def reset_settings(self):
        """
        Reload all settings from config.json.
        """

        # Reload config
        self.config_obj = load_config(self.config_path)
        self.base_boxes = deepcopy(self.config_obj.boxes)

        q = self.config_obj.get_style("question")
        a = self.config_obj.get_style("answer")

        # Font
        self.font_var.set(self.config_obj.get_font_family())

        self.q_max.set(q.get("font_max", 32))
        self.q_min.set(q.get("font_min", 18))
        self.a_max.set(a.get("font_max", 28))
        self.a_min.set(a.get("font_min", 20))

        # Offsets
        self.q_offset.set(0)
        self.a_offset.set(0)

        # Question Box
        self.q_left.set(0)
        self.q_right.set(0)
        self.q_top.set(0)
        self.q_bottom.set(0)

        # Answer Box
        self.a_left.set(0)
        self.a_right.set(0)
        self.a_top.set(0)
        self.a_bottom.set(0)

        # Margins
        self.q_margin_left.set(q.get("margin_left", 20))
        self.q_margin_right.set(q.get("margin_right", 20))
        self.q_margin_top.set(q.get("margin_top", 10))
        self.q_margin_bottom.set(q.get("margin_bottom", 10))

        self.a_margin_left.set(a.get("margin_left", 20))
        self.a_margin_right.set(a.get("margin_right", 20))
        self.a_margin_top.set(a.get("margin_top", 10))
        self.a_margin_bottom.set(a.get("margin_bottom", 10))

        # Colors
        self.green_q_fill = "#FFFFFF"
        self.green_a_fill = "#FFFFFF"
        self.green_q_outline = "#000000"
        self.green_a_outline = "#000000"

        self.red_q_fill = "#FFFF00"
        self.red_a_fill = "#FFFF00"
        self.red_q_outline = "#000000"
        self.red_a_outline = "#000000"

        # Subject & Stage
        self.green_subject_fill="#FFFFFF"
        self.green_subject_outline="#000000"
        self.green_stage_fill="#FFFFFF"
        self.green_stage_outline="#000000"

        self.red_subject_fill="#FFFF00"
        self.red_subject_outline="#000000"
        self.red_stage_fill="#FFFF00"
        self.red_stage_outline="#000000"

        # ---------------- Clear Project Selection ----------------
        self.green_template_path = None
        self.red_template_path = None

        self.excel_path = None
        self.output_dir = None

        self.start_row.set(2)
        self.end_row.set(0)

        # Clear previews
        self.green_preview.configure(image="")
        self.red_preview.configure(image="")

        self.green_photo = None
        self.red_photo = None

        # Clear log
        self.log_box.delete("1.0", tk.END)

        # Reset status
        self.status.set("Ready")

        self.update_preview()

        self.status.set("Settings reset.")

        self.green_template_label.config(text="Not Selected", foreground="gray")
        self.red_template_label.config(text="Not Selected", foreground="gray")
        self.excel_label.config(text="Not Selected", foreground="gray")
        self.output_label.config(text="Not Selected", foreground="gray")


    def load_data_file(self, filename):

        import os
        import csv

        ext = os.path.splitext(filename)[1].lower()

        rows = []

        # ---------------- Excel ----------------

        if ext == ".xlsx":

            ##from openpyxl import load_workbook##

            wb = load_workbook(filename, data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):

                rows.append(
                    (
                        row[0],
                        row[1],
                        row[2]
                    )
                )

            return rows

        # ---------------- CSV ----------------

        elif ext == ".csv":

            with open(
                filename,
                newline="",
                encoding="utf-8-sig"
            ) as f:

                reader = csv.reader(f)

                next(reader, None)

                for row in reader:

                    rows.append(
                        (
                            row[0],
                            row[1],
                            row[2]
                        )
                    )

            return rows

        # ---------------- TXT ----------------

        elif ext == ".txt":

            with open(
                filename,
                encoding="utf-8"
            ) as f:

                first = f.readline()

                if "\t" in first:
                    delimiter = "\t"

                elif "|" in first:
                    delimiter = "|"

                else:
                    delimiter = ","

                f.seek(0)

                reader = csv.reader(
                    f,
                    delimiter=delimiter
                )

                next(reader, None)

                for row in reader:

                    rows.append(
                        (
                            row[0],
                            row[1],
                            row[2]
                        )
                    )

            return rows

        # ---------------- DOCX ----------------

        elif ext == ".docx":

            ##from docx import Document##

            doc = Document(filename)

            table = doc.tables[0]

            for row in table.rows[1:]:

                cells = row.cells

                rows.append(
                    (
                        cells[0].text,
                        cells[1].text,
                        cells[2].text
                    )
                )

            return rows

        else:

            raise Exception(
                "Unsupported file format."
            )

    # =====================================================
    # GENERATION
    # =====================================================
    def start_generation(self):
        self.status.set("Generating cards...")
        threading.Thread(target=self.generate_cards, daemon=True).start()

    def generate_cards(self):
        # Excel and output are always required
        if not self.excel_path or not self.output_dir:
            self.root.after(
                0,
                lambda: messagebox.showerror(
                    "Error",
                    "Please select a data file and an Output folder."
                )
            )
            return

        # At least one template is required
        if not self.green_template_path and not self.red_template_path:
            self.root.after(
                0,
                lambda: messagebox.showerror(
                    "Error",
                    "Please select at least one card template."
                )
            )
            return

        rows = self.load_data_file(self.excel_path)

        if not rows:
            self.root.after(
                0,
                lambda: messagebox.showerror(
                    "Error",
                    "No valid data found in the selected file."
                )
            )
            return

        config = deepcopy(self.config_obj)

        q = self.base_boxes["question"]
        a = self.base_boxes["answer"]

        config.boxes["question"] = [
            q[0] + self.q_left.get(),
            q[1] + self.q_top.get() + self.q_offset.get(),
            q[2] + self.q_right.get(),
            q[3] + self.q_bottom.get(),
        ]

        config.boxes["answer"] = [
            a[0] + self.a_left.get(),
            a[1] + self.a_offset.get() + self.a_top.get(),
            a[2] + self.a_right.get(),
            a[3] + self.a_bottom.get(),
        ]

        subject_box = self.base_boxes["subject"]

        config.boxes["subject"] = [
            subject_box[0] - self.subject_left.get() + self.subject_right.get(),
            subject_box[1] - self.subject_up.get() + self.subject_down.get(),
            subject_box[2],
            subject_box[3],
        ]

        stage_box = self.base_boxes["stage"]

        config.boxes["stage"] = [
            stage_box[0] - self.stage_left.get() + self.stage_right.get(),
            stage_box[1] - self.stage_up.get() + self.stage_down.get(),
            stage_box[2],
            stage_box[3],
        ]

        # Fonts
        config.style["question"]["font_max"] = self.q_max.get()
        config.style["question"]["font_min"] = self.q_min.get()

        config.style["answer"]["font_max"] = self.a_max.get()
        config.style["answer"]["font_min"] = self.a_min.get()

        # Margins
        config.style["question"]["margin_left"] = self.q_margin_left.get()
        config.style["question"]["margin_right"] = self.q_margin_right.get()
        config.style["question"]["margin_top"] = self.q_margin_top.get()
        config.style["question"]["margin_bottom"] = self.q_margin_bottom.get()

        config.style["answer"]["margin_left"] = self.a_margin_left.get()
        config.style["answer"]["margin_right"] = self.a_margin_right.get()
        config.style["answer"]["margin_top"] = self.a_margin_top.get()
        config.style["answer"]["margin_bottom"] = self.a_margin_bottom.get()

        self.output_dir.mkdir(exist_ok=True)

        # ---------------- Subject ----------------

        config.style["subject"]["font_size"] = self.subject_font_size.get()
        config.style["subject"]["bold"] = self.subject_bold.get()
        config.style["subject"]["italic"] = self.subject_italic.get()

        # ---------------- Stage ----------------

        config.style["stage"]["font_size"] = self.stage_font_size.get()
        config.style["stage"]["bold"] = self.stage_bold.get()
        config.style["stage"]["italic"] = self.stage_italic.get()

        start = self.start_row.get()
        end = self.end_row.get()

        if end <= 0:
            end = len(rows)
        
        if start < 1:
            self.root.after(
                0,
                lambda: messagebox.showerror(
                    "Error",
                    "Start Row must be 1 or greater."
                )
            )
            return

        if start > end:
            self.root.after(
                0,
                lambda: messagebox.showerror(
                    "Error",
                    "Start Row cannot be greater than End Row."
                )
            )
            return

        if end > len(rows):
            end = len(rows)
        
        card_no = 1
        
        for i, row in enumerate(rows[start-1:end], start):

            if not row:
                continue

            # Need at least Question and Answer
            if len(row) < 2:
                self.log(f"Skipped Row {i} (Not enough columns)")
                continue

            if not row[0] or not row[1]:
                continue

            # Read Points safely
            try:
                points = int(row[2]) if len(row) >= 3 and row[2] is not None else 0
            except (TypeError, ValueError):
                points = 0

            if points == 10:

                if not self.green_template_path:
                    self.log(f"Skipped Row {i} (Green template not selected)")
                    continue

                cfg = deepcopy(config)

                cfg.style["question"]["fill"] = self.green_q_fill
                cfg.style["answer"]["fill"] = self.green_a_fill
                cfg.style["question"]["stroke_fill"] = self.green_q_outline
                cfg.style["answer"]["stroke_fill"] = self.green_a_outline
                cfg.style["subject"]["fill"] = self.green_subject_fill
                cfg.style["subject"]["stroke_fill"] = self.green_subject_outline
                cfg.style["stage"]["fill"] = self.green_stage_fill
                cfg.style["stage"]["stroke_fill"] = self.green_stage_outline

                template = self.green_template_path


            elif points == 20:

                if not self.red_template_path:
                    self.log(f"Skipped Row {i} (Red template not selected)")
                    continue

                cfg = deepcopy(config)

                cfg.style["question"]["fill"] = self.red_q_fill
                cfg.style["answer"]["fill"] = self.red_a_fill
                cfg.style["question"]["stroke_fill"] = self.red_q_outline
                cfg.style["answer"]["stroke_fill"] = self.red_a_outline
                cfg.style["subject"]["fill"] = self.red_subject_fill
                cfg.style["subject"]["stroke_fill"] = self.red_subject_outline
                cfg.style["stage"]["fill"] = self.red_stage_fill
                cfg.style["stage"]["stroke_fill"] = self.red_stage_outline

                template = self.red_template_path


            else:

                self.log(f"Skipped Row {i} (Invalid Points = {points})")
                continue


            create_card(
                question=str(row[0]),
                answer=str(row[1]),

                subject=self.subject_var.get(),
                stage=self.stage_var.get(),

                subject_font_path=self.font_map.get(
                    self.subject_font_var.get(),
                    "Arial"
                ),

                stage_font_path=self.font_map.get(
                    self.stage_font_var.get(),
                    "Arial"
                ),

                output_file=self.output_dir / f"Card_{card_no:03}.png",
                config=cfg,
                template_path=template,
                font_path=self.font_map.get(
                    self.font_var.get(),
                    "Arial"
                ),
            )

            card_no += 1

            self.log(f"Generated Card_{card_no-1:03}.png")

            self.root.after(
                0,
                lambda i=i: self.status.set(f"Generating Card_{card_no-1:03}.png")
            )

        
        self.root.after(
            0,
            lambda: (
                self.status.set("Generation complete"),
                messagebox.showinfo("Done", "Generation complete")
            )
        )

    # =====================================================
    # LOGGING
    # =====================================================
    def show_about(self):

        import tkinter as tk
        from tkinter import ttk

        from src.version import APP_VERSION

        win = tk.Toplevel(self.root)

        win.title("About Trump Card Generator")
        win.geometry("500x430")
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        try:
            win.iconbitmap("icon.ico")
        except:
            pass

        frame = ttk.Frame(
            win,
            padding=20
        )

        frame.pack(
            fill="both",
            expand=True
        )

        ttk.Label(
            frame,
            text="Trump Card Generator",
            font=("Segoe UI", 16, "bold")
        ).pack()

        ttk.Label(
            frame,
            text=f"Version {APP_VERSION}",
            foreground="blue",
            font=("Segoe UI", 10)
        ).pack(pady=(5,15))

        ttk.Label(
            frame,
            text="Developed by",
            font=("Segoe UI",10,"bold")
        ).pack()

        ttk.Label(
            frame,
            text="Cogni Cards"
        ).pack(pady=(0,15))

        ttk.Label(
            frame,
            text="Create printable Trump Cards directly\nfrom Excel in just a few clicks.",
            justify="center"
        ).pack()

        ttk.Separator(frame).pack(
            fill="x",
            pady=15
        )

        ttk.Label(
            frame,
            text="Features",
            font=("Segoe UI",10,"bold")
        ).pack(anchor="w")

        features = [
            "✔ Green Card Generation",
            "✔ Red Card Generation",
            "✔ Live Preview",
            "✔ Excel Import",
            "✔ Subject & Stage Support",
            "✔ Auto Update",
            "✔ High Resolution Export"
        ]

        for item in features:

            ttk.Label(
                frame,
                text=item
            ).pack(anchor="w")

        ttk.Separator(frame).pack(
            fill="x",
            pady=15
        )

        ttk.Label(
            frame,
            text="GitHub",
            font=("Segoe UI",10,"bold")
        ).pack()

        github = ttk.Label(
            frame,
            text="https://github.com/UjjalKrRoy",
            foreground="blue",
            cursor="hand2"
        )

        github.pack()

        github.bind(
            "<Button-1>",
            lambda e: webbrowser.open(
                "https://github.com/UjjalKrRoy"
            )
        )

        ttk.Label(
            frame,
            text="© 2026 Ujjal Roy",
            foreground="gray"
        ).pack(
            pady=(15,10)
        )

        btn_frame = ttk.Frame(frame)
        btn_frame.pack()

        ttk.Button(
            btn_frame,
            text="Visit GitHub",
            command=lambda: webbrowser.open(
                "https://github.com/UjjalKrRoy"
            )
        ).pack(
            side="left",
            padx=5
        )

        ttk.Button(
            btn_frame,
            text="Close",
            command=win.destroy
        ).pack(
            side="left",
            padx=5
        )
    def log(self, msg):
        self.root.after(
            0,
            lambda: (
                self.log_box.insert("end", msg + "\n"),
                self.log_box.see("end")
            )
        )
    def check_updates(self):

        update_available, version = check_for_update()

        if not update_available:

            messagebox.showinfo(
                "Software Update",
                "You are already using the latest version."
            )
            return

        answer = messagebox.askyesno(
            "Software Update",
            f"Version {version} is available.\n\n"
            "Do you want to update now?"
        )

        if answer:
            download_and_replace()


if __name__ == "__main__":
    root = tk.Tk()
    CardGeneratorGUI(root)
    root.mainloop()